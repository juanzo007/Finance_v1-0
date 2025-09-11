#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
finances_pipeline.py
- OCR images in data/test-images
- Extract Date/Time (date_extractor.py)
- Extract THB Withdrawal (amount_extractor.py)
- Extract Description (description_extractor.py) and Note (note_extractor.py)
- Write to Finances.xlsx in project root
- Finalize numbers to 2 decimals
- Store bare filename; add "Open" hyperlink to image
"""

from __future__ import annotations

import importlib.util
import os
import sys
from pathlib import Path
from typing import List, Tuple, Dict, Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------- OCR libs ----------
try:
    from PIL import Image, ImageOps, ImageFilter
except Exception:
    Image = None
    ImageOps = None
    ImageFilter = None

try:
    import pytesseract
except Exception:
    pytesseract = None

# ---------- Paths ----------
HERE = Path(__file__).resolve().parent
IMG_DIR = HERE / "data" / "test-images"
EXTRACTOR_DATE_PATH = HERE / "scripts" / "image-scipts" / "date_extractor.py"
EXTRACTOR_AMOUNT_PATH = HERE / "scripts" / "image-scipts" / "amount_extractor.py"
EXTRACTOR_DESC_PATH = HERE / "scripts" / "image-scipts" / "description_extractor.py"
EXTRACTOR_NOTE_PATH = HERE / "scripts" / "image-scipts" / "note_extractor.py"
OUT_XLSX = HERE / "Finances.xlsx"

HEADERS = [
    "Date",
    "Time",
    "THB Withdrawal",
    "THB Deposit",
    "USD Amount",
    "FX rate",
    "Description",
    "Acct #",
    "Merchant_ID",
    "Note",
    "Sub_Category",
    "Category",
    "Filename",
    "Source",
    "Open File",
]
TEXT_COLS = ["Description", "Note", "Filename", "Source", "Open File"]


# ---------- Utils ----------
def _log(msg: str):
    print(msg)


def _ensure_tesseract():
    if pytesseract is None:
        return
    if os.name == "nt":
        guess = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(guess):
            pytesseract.pytesseract.tesseract_cmd = guess


def _ocr_image(img_path: Path) -> Tuple[str, List[str]]:
    """Return (full_text, lines). Always returns safely; never raises."""
    if pytesseract is None or Image is None:
        _log(f"[WARN] OCR libs missing for {img_path.name}; skipping OCR.")
        return "", []

    try:
        img = Image.open(img_path)
    except Exception as e:
        _log(f"[WARN] Cannot open image {img_path.name}: {e}")
        return "", []

    # Pre-process (best-effort)
    try:
        img = img.convert("L")
    except Exception:
        pass
    try:
        img = ImageOps.autocontrast(img)
    except Exception:
        pass
    try:
        img = img.filter(ImageFilter.MedianFilter(size=3))
    except Exception:
        pass  # ignore rankfilter quirks

    # Language
    try:
        langs = set(pytesseract.get_languages(config=""))
        lang = "eng+tha" if "tha" in langs else "eng"
    except Exception:
        lang = "eng"

    # Timed OCR
    try:
        full = pytesseract.image_to_string(img, lang=lang, timeout=12)
    except RuntimeError as e:
        _log(f"[WARN] Tesseract error on {img_path.name}: {e}")
        full = ""
    except Exception as e:
        _log(f"[WARN] OCR failed on {img_path.name}: {e}")
        full = ""

    lines: List[str] = (
        [ln.strip() for ln in full.splitlines() if ln.strip()] if full.strip() else []
    )

    # Structured lines (timed)
    try:
        df = pytesseract.image_to_data(
            img, lang=lang, output_type=pytesseract.Output.DATAFRAME, timeout=12
        )
        if df is not None:
            df = df.dropna(subset=["text"])
            structured = []
            for ln in sorted(df.get("line", []).dropna().unique()):
                row = df[df["line"] == ln]
                s = " ".join(str(t).strip() for t in row["text"] if str(t).strip())
                if s:
                    structured.append(s)
            if structured:
                lines = structured
    except RuntimeError as e:
        _log(f"[WARN] Tesseract data error on {img_path.name}: {e}")
    except Exception:
        pass

    _log(f"[OCR] {img_path.name}: {len(lines)} line(s) recognized.")
    return full, lines


def _dynamic_import(py_path: Path):
    if not py_path.exists():
        _log(f"[ERROR] Extractor not found: {py_path}")
        return None
    try:
        spec = importlib.util.spec_from_file_location(py_path.stem, str(py_path))
        if not spec or not spec.loader:
            return None
        mod = importlib.util.module_from_spec(spec)
        sys.modules[py_path.stem] = mod
        spec.loader.exec_module(mod)  # type: ignore[attr-defined]
        return mod
    except Exception as e:
        _log(f"[ERROR] Could not import extractor {py_path.name}: {e}")
        return None


def _ensure_headers(df: pd.DataFrame) -> pd.DataFrame:
    for h in HEADERS:
        if h not in df.columns:
            df[h] = ""
    for col in TEXT_COLS:
        if col in df.columns:
            df[col] = df[col].astype("object")
    return df[HEADERS]


def _read_or_new() -> pd.DataFrame:
    if OUT_XLSX.exists():
        df = pd.read_excel(OUT_XLSX, engine="openpyxl")
        return _ensure_headers(df)
    return _ensure_headers(pd.DataFrame(columns=HEADERS))


def _finalize_numbers(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["THB Withdrawal", "THB Deposit", "USD Amount", "FX rate"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(2)
    return df


def _save_with_excel_format(df: pd.DataFrame, path: Path):
    """Save DataFrame with 2-decimal format + clickable 'Open' hyperlinks."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]

        # Amount columns: enforce 2 decimals
        for col_name in ["THB Withdrawal", "THB Deposit", "USD Amount"]:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"

        # Hyperlinks: show 'Open' text, link to absolute path
        if "Open File" in df.columns and "Filename" in df.columns:
            open_idx = df.columns.get_loc("Open File") + 1
            fn_idx = df.columns.get_loc("Filename") + 1
            for row in range(2, ws.max_row + 1):
                fn_val = ws.cell(row=row, column=fn_idx).value
                if fn_val:
                    img_abspath = (IMG_DIR / fn_val).resolve()
                    cell = ws.cell(row=row, column=open_idx)
                    cell.value = "Open"
                    cell.hyperlink = img_abspath.as_uri()  # guaranteed-open file:// URI
                    cell.font = Font(color="0000EE", underline="single")


def _upsert(filename_only: str, payload: Dict[str, Any]):
    df = _read_or_new()

    if (df["Filename"] == filename_only).any():
        idx = df.index[df["Filename"] == filename_only][0]
        for k, v in payload.items():
            if v not in ("", None):
                df.at[idx, k] = v
    else:
        row = {h: "" for h in HEADERS}
        row.update(
            {
                "Filename": filename_only,
                "Open File": filename_only,
                "Source": "Bangkok Bank Receipt",
            }
        )
        for k, v in payload.items():
            if v not in ("", None):
                row[k] = v
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    df = _finalize_numbers(df)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    _save_with_excel_format(df, OUT_XLSX)


# ---------- Main ----------
def main():
    _ensure_tesseract()

    # Import extractors
    ext_date = _dynamic_import(EXTRACTOR_DATE_PATH)
    ext_amount = _dynamic_import(EXTRACTOR_AMOUNT_PATH)
    ext_desc = _dynamic_import(EXTRACTOR_DESC_PATH)
    ext_note = _dynamic_import(EXTRACTOR_NOTE_PATH)

    img_paths = (
        [
            p
            for p in IMG_DIR.glob("*.*")
            if p.suffix.lower() in {".png", ".jpg", ".jpeg"}
        ]
        if IMG_DIR.exists()
        else []
    )
    _log(f"[INFO] Found {len(img_paths)} image(s) in {IMG_DIR}")

    for img in img_paths:
        text, lines = _ocr_image(img)

        date_val, time_val, thb_withdrawal = "", "", None
        description, note = "", ""

        # ---- Date/Time via date_extractor ----
        if ext_date and hasattr(ext_date, "extract_date_time"):
            try:
                out = ext_date.extract_date_time(
                    str(img), text, lines
                )  # {'date','time'}
                if isinstance(out, dict):
                    date_val = (out.get("date") or "").strip()
                    time_val = (out.get("time") or "").strip()
            except Exception as e:
                _log(f"[WARN] date extract failed for {img.name}: {e}")

        # ---- Amount via amount_extractor ----
        if ext_amount and hasattr(ext_amount, "extract"):
            try:
                out_amt = ext_amount.extract(
                    str(img), text, lines
                )  # {'thb_withdrawal': float or ""}
                if isinstance(out_amt, dict):
                    val = out_amt.get("thb_withdrawal", "")
                    if val not in ("", None):
                        thb_withdrawal = round(float(val), 2)
            except Exception as e:
                _log(f"[WARN] amount extract failed for {img.name}: {e}")

        # ---- Description / Note via their extractors ----
        if ext_desc and hasattr(ext_desc, "extract"):
            try:
                od = ext_desc.extract(str(img), text, lines)
                if isinstance(od, dict):
                    description = (od.get("description") or "").strip()
            except Exception as e:
                _log(f"[WARN] description extract failed for {img.name}: {e}")

        if ext_note and hasattr(ext_note, "extract"):
            try:
                on = ext_note.extract(str(img), text, lines)
                if isinstance(on, dict):
                    note = (on.get("note") or "").strip()
            except Exception as e:
                _log(f"[WARN] note extract failed for {img.name}: {e}")

        # Sanity filter: drop classic false hits
        if description.strip().lower() in {"verify", "veri"}:
            description = ""

        payload = {
            "Date": date_val,
            "Time": time_val,
            "THB Withdrawal": thb_withdrawal,
            "Description": description,
            "Note": note,
        }
        _upsert(img.name, payload)

        desc_preview = (
            (description[:40] + "…")
            if description and len(description) > 40
            else (description or "(none)")
        )
        note_preview = (
            (note[:40] + "…") if note and len(note) > 40 else (note or "(none)")
        )

        _log(
            f"[OK] {img.name}\n"
            f"   Date: {date_val or '(none)'}   Time: {time_val or '(none)'}\n"
            f"   THB Withdrawal: {(f'{thb_withdrawal:,.2f}' if thb_withdrawal is not None else '(none)')}\n"
            f"   Description: {desc_preview}\n"
            f"   Note: {note_preview}"
        )

    _log(f"[DONE] Updated: {OUT_XLSX}")


if __name__ == "__main__":
    main()
