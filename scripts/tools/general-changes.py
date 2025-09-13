# -*- coding: utf-8 -*-
"""
general-changes.py - Post-processing utilities for the finance pipeline
Place this in /scripts/general-changes.py

Usage: python scripts/general-changes.py
"""

import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

# Paths
ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_FILE = ROOT / "Finances.xlsx"
IMAGE_DIR = ROOT / "data" / "bangkok_images"


def convert_image_column_to_hyperlinks():
    """Convert Image column values to hyperlinks pointing to actual image files"""
    print(f"[INFO] Processing {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel file not found: {EXCEL_FILE}")
        return

    # First, let's see what's in the Excel file
    try:
        import pandas as pd

        df = pd.read_excel(EXCEL_FILE)
        print(f"[DEBUG] Excel columns: {list(df.columns)}")
        print(f"[DEBUG] Excel shape: {df.shape}")

        if "Image" in df.columns:
            print(f"[DEBUG] Image column values:")
            for idx, val in df["Image"].head(10).items():
                print(f"  Row {idx}: '{val}'")
        else:
            print("[ERROR] No 'Image' column found!")
            return
    except Exception as e:
        print(f"[ERROR] Could not read Excel with pandas: {e}")
        return

    # Load the Excel file with openpyxl for hyperlink support
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"[DEBUG] Workbook loaded, active sheet: {ws.title}")
    except Exception as e:
        print(f"[ERROR] Could not load workbook: {e}")
        return

    # Find the Image column
    image_col_idx = None
    header_row = 1

    print("[DEBUG] Checking headers in row 1:")
    for col_idx, cell in enumerate(ws[header_row], 1):
        print(f"  Column {col_idx}: '{cell.value}'")
        if cell.value == "Image":
            image_col_idx = col_idx
            break

    if image_col_idx is None:
        print("[ERROR] No 'Image' column found in Excel file")
        return

    print(f"[INFO] Found Image column at index {image_col_idx}")

    # Process each row (skip header)
    updated_count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=image_col_idx)

        if cell.value and isinstance(cell.value, str):
            filename = cell.value.strip()
            print(f"[DEBUG] Processing row {row_idx}: '{filename}'")

            # Create the relative path from Excel file location
            image_path = f"data\\bangkok_images\\{filename}"

            # Check if file exists
            full_path = IMAGE_DIR / filename
            if full_path.exists():
                # Create hyperlink
                cell.hyperlink = image_path
                cell.value = filename  # Keep the filename as display text
                cell.style = "Hyperlink"  # Apply Excel hyperlink style
                updated_count += 1
                print(f"[OK] Row {row_idx}: {filename} -> {image_path}")
            else:
                print(f"[WARN] Row {row_idx}: File not found - {full_path}")
        else:
            print(f"[DEBUG] Row {row_idx}: Empty or invalid cell value: {cell.value}")

    # Save the workbook
    try:
        wb.save(EXCEL_FILE)
        print(f"[DONE] Updated {updated_count} hyperlinks in {EXCEL_FILE}")
    except Exception as e:
        print(f"[ERROR] Could not save workbook: {e}")


def remove_image_hyperlinks():
    """Remove hyperlinks from Image column (keep just the filenames)"""
    print(f"[INFO] Removing hyperlinks from {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel file not found: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find the Image column
    image_col_idx = None
    header_row = 1

    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value == "Image":
            image_col_idx = col_idx
            break

    if image_col_idx is None:
        print("[WARN] No 'Image' column found in Excel file")
        return

    # Remove hyperlinks from each row
    updated_count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=image_col_idx)

        if cell.hyperlink:
            cell.hyperlink = None
            cell.style = "Normal"
            updated_count += 1

    wb.save(EXCEL_FILE)
    print(f"[DONE] Removed {updated_count} hyperlinks from {EXCEL_FILE}")


def validate_image_files():
    """Check which image files in the Excel exist in the bangkok_images directory"""
    print(f"[INFO] Validating image files referenced in {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        print(f"[ERROR] Excel file not found: {EXCEL_FILE}")
        return

    df = pd.read_excel(EXCEL_FILE)

    if "Image" not in df.columns:
        print("[WARN] No 'Image' column found")
        return

    missing_files = []
    existing_files = []

    for idx, row in df.iterrows():
        if pd.notna(row["Image"]):
            filename = str(row["Image"]).strip()
            file_path = IMAGE_DIR / filename

            if file_path.exists():
                existing_files.append(filename)
            else:
                missing_files.append(filename)

    print(f"[INFO] Files found: {len(existing_files)}")
    print(f"[WARN] Files missing: {len(missing_files)}")

    if missing_files:
        print("[MISSING FILES]")
        for f in missing_files:
            print(f"  - {f}")

    return existing_files, missing_files


def main():
    """Run the post-processing operations"""
    print("=== Finance Pipeline Post-Processing ===")

    # You can uncomment the operations you want to run:

    # 1. Convert Image column to hyperlinks
    convert_image_column_to_hyperlinks()

    # 2. Validate that image files exist
    # validate_image_files()

    # 3. Remove hyperlinks (if needed)
    # remove_image_hyperlinks()


if __name__ == "__main__":
    main()
